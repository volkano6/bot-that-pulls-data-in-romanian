import time

from selenium import webdriver
from selenium.webdriver.common.by import By

from openpyxl.workbook import Workbook
from openpyxl import load_workbook

driver = webdriver.Chrome()
url = "https://www.news.ro/sport/"

#Create a workbook obj
#wb = Workbook()

#load existing spredsheet
wb = load_workbook(r'C:\Users\MONSTER\Desktop\output.xlsx')

#create an active worksheet
ws = wb.active


def main():
    news = []
    sentence_count = 2400
    loop_is_continue = True
    main_page = 1

    driver.get(url)
    time.sleep(3)

    aaa = driver.find_element(By.CSS_SELECTOR, "#onetrust-accept-btn-handler")
    aaa.click()
    print("sekme kapandı")

    time.sleep(3)

    while loop_is_continue:

        print(f"Main Page: {main_page}")

        news = driver.find_elements(By.CSS_SELECTOR, ".row.article")
        for new in range(len(news)):

            sentence = ""

            time.sleep(1)
            driver.find_element(By.XPATH, f"/html/body/div[7]/div/main/section/article[{new + 1}]/div[2]/h2/a").click()
            print(f"page: {new+1}")
            time.sleep(2)

            articles = driver.find_elements(By.TAG_NAME, 'p')

            for i in range(len(articles)):
                sentence += articles[i].text

            sentence = sentence.replace("Articolul de mai sus este destinat exclusiv informării dumneavoastră personale. Dacă reprezentaţi o instituţie media sau o companie şi doriţi un acord pentru republicarea articolelor noastre, va rugăm să ne trimiteţi un mail pe adresa abonamente@news.ro.Înscrie-te la Newsletter-ul zilnic News.ro și vei primi cele mai importante știri în fiecare dimineață pe adresa ta de email.Nu îți face griji, nu te spamăm.Te poți dezabona cu un singur click.© 2022 News.ro. Toate drepturile rezervate.Dezvoltat de 1616.ro","")
            sentence = sentence.replace("  ", "")
            sentence = sentence.replace("•", " ")
            ws[f"A{sentence_count+1}"].value = sentence
            wb.save(r'C:\Users\MONSTER\Desktop\output.xlsx')

            sentence_count += 1

            driver.back()
            print(sentence_count)
            print("------------")

        driver.find_element(By.XPATH, "/html/body/div[7]/div/main/nav/ul/li[7]/a").click()
        time.sleep(4)

        if sentence_count >= 3000:
            loop_is_continue = False

        main_page += 1
        print("////////////////")

    driver.close()
    print("sayfa kapandı")


if __name__ == '__main__':
    main()
